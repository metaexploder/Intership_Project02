"""
PACE Payroll Validator — Report Generator
============================================
Produces the final Excel validation report using openpyxl.

Columns:
    WOID | Coverage | Sheets | Files | InceptionDate | ExpirationDate |
    PeriodStart | PeriodEnd | StartBuffer | EndBuffer
"""

from datetime import date
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import BUFFER_DAYS
from logger_setup import setup_logger

logger = setup_logger()

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
_HEADER_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_TITLE_FONT   = Font(name="Calibri", bold=True, size=14, color="1F4E79")
_METRIC_LABEL = Font(name="Calibri", bold=True, size=11)
_METRIC_VALUE = Font(name="Calibri", size=11)
_DATA_FONT    = Font(name="Calibri", size=10)

_STATUS_FONTS = {
    "FULL":    Font(name="Calibri", bold=True, color="217346"),   # green
    "PARTIAL": Font(name="Calibri", bold=True, color="BF8F00"),   # amber
    "NO":      Font(name="Calibri", bold=True, color="C00000"),   # red
}
_STATUS_FILLS = {
    "FULL":    PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "PARTIAL": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "NO":      PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
}

# Buffer cell coloring: green if surplus (>= 0), red if deficit (< 0)
_BUF_OK_FONT  = Font(name="Calibri", size=10, color="217346", bold=True)
_BUF_BAD_FONT = Font(name="Calibri", size=10, color="C00000", bold=True)

_THIN_BORDER = Border(
    left=Side(style="thin",  color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin",   color="B0B0B0"),
    bottom=Side(style="thin",color="B0B0B0"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT   = Alignment(horizontal="left",   vertical="center")

# Column definitions: (header label, width)
_COLUMNS = [
    ("WOID",           18),
    ("Coverage",       14),
    ("Sheets",         12),
    ("Files",          12),
    ("InceptionDate",  16),
    ("ExpirationDate", 16),
    ("PeriodStart",    15),
    ("PeriodEnd",      15),
    (f"StartBuffer\n(±{BUFFER_DAYS}d)", 16),
    (f"EndBuffer\n(±{BUFFER_DAYS}d)",   16),
]

# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_report(results: List[Dict], output_path: Path) -> None:
    """
    Generate ``PACE_PayPeriod_Validation_Report.xlsx``.

    Parameters
    ----------
    results : List[Dict]
        Each dict contains:
            - woid            : str
            - status          : str ("FULL" / "PARTIAL" / "NO")
            - missing_gaps    : List[Tuple[date, date]]
            - sheets_processed: int
            - files_count     : int
            - policy_start    : Optional[date]
            - policy_end      : Optional[date]
            - period_start    : Optional[date]  (min across all sheets/files)
            - period_end      : Optional[date]  (max across all sheets/files)
            - start_buffer    : Optional[int]   (days payroll starts before inception)
            - end_buffer      : Optional[int]   (days payroll ends after expiration)
    output_path : Path
        Destination file path for the report.
    """
    wb = Workbook()
    _build_summary_sheet(wb, results)
    wb.save(str(output_path))
    logger.info("Report saved -> %s", output_path)


# ---------------------------------------------------------------------------
# Sheet builder
# ---------------------------------------------------------------------------

def _build_summary_sheet(wb: Workbook, results: List[Dict]) -> None:
    """Create the Summary worksheet with all 10 columns."""
    ws = wb.active
    ws.title = "Summary"

    # --- Set column widths ---
    for col_idx, (_, width) in enumerate(_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Counts ---
    total         = len(results)
    full_count    = sum(1 for r in results if r["status"] == "FULL")
    partial_count = sum(1 for r in results if r["status"] == "PARTIAL")
    no_count      = sum(1 for r in results if r["status"] == "NO")
    num_cols      = len(_COLUMNS)

    # Title row (spans all columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws["A1"]
    title_cell.value     = "PayPeriod Validation Summary"
    title_cell.font      = _TITLE_FONT
    title_cell.alignment = _LEFT
    ws.row_dimensions[1].height = 22

    # Metrics block (rows 3-6)
    metrics = [
        ("Total WOIDs Processed:", total),
        ("FULL:",                  full_count),
        ("PARTIAL:",               partial_count),
        ("NO:",                    no_count),
    ]
    for idx, (label, value) in enumerate(metrics, start=3):
        lbl = ws.cell(row=idx, column=1, value=label)
        lbl.font = _METRIC_LABEL
        val = ws.cell(row=idx, column=2, value=value)
        val.font      = _METRIC_VALUE
        val.alignment = _CENTER

    # --- Header row ---
    header_row = 8
    ws.row_dimensions[header_row].height = 30
    for col_idx, (header, _) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _THIN_BORDER

    # --- Data rows ---
    for row_offset, res in enumerate(results, start=1):
        row_num = header_row + row_offset
        status  = res["status"]

        def _cell(col: int, value, fmt: str = None):
            c = ws.cell(row=row_num, column=col, value=value)
            c.font      = _DATA_FONT
            c.alignment = _CENTER
            c.border    = _THIN_BORDER
            if fmt:
                c.number_format = fmt
            return c

        def _fmt_date(d: Optional[date]) -> Optional[str]:
            return d.strftime("%Y-%m-%d") if d else "N/A"

        def _fmt_buf(val: Optional[int]) -> str:
            if val is None:
                return "N/A"
            return f"+{val}d" if val >= 0 else f"{val}d"

        # Col 1: WOID
        _cell(1, res["woid"])

        # Col 2: Coverage (color-coded)
        cov = ws.cell(row=row_num, column=2, value=status)
        cov.font      = _STATUS_FONTS.get(status, _DATA_FONT)
        cov.fill      = _STATUS_FILLS.get(status, PatternFill())
        cov.alignment = _CENTER
        cov.border    = _THIN_BORDER

        # Col 3: Sheets
        _cell(3, "Multiple" if res.get("sheets_processed", 1) > 1 else "Single")

        # Col 4: Files
        _cell(4, "Multiple" if res.get("files_count", 1) > 1 else "Single")

        # Col 5: InceptionDate
        _cell(5, _fmt_date(res.get("policy_start")))

        # Col 6: ExpirationDate
        _cell(6, _fmt_date(res.get("policy_end")))

        # Col 7: PeriodStart (min start across all payroll sheets/files)
        _cell(7, _fmt_date(res.get("period_start")))

        # Col 8: PeriodEnd (max end across all payroll sheets/files)
        _cell(8, _fmt_date(res.get("period_end")))

        # Col 9: StartBuffer (days payroll starts before inception; positive = surplus)
        sb_val = res.get("start_buffer")
        sb_cell = ws.cell(row=row_num, column=9, value=_fmt_buf(sb_val))
        sb_cell.font      = (_BUF_OK_FONT if sb_val is not None and sb_val >= 0 else _BUF_BAD_FONT) if sb_val is not None else _DATA_FONT
        sb_cell.alignment = _CENTER
        sb_cell.border    = _THIN_BORDER

        # Col 10: EndBuffer (days payroll ends after expiration; positive = surplus)
        eb_val = res.get("end_buffer")
        eb_cell = ws.cell(row=row_num, column=10, value=_fmt_buf(eb_val))
        eb_cell.font      = (_BUF_OK_FONT if eb_val is not None and eb_val >= 0 else _BUF_BAD_FONT) if eb_val is not None else _DATA_FONT
        eb_cell.alignment = _CENTER
        eb_cell.border    = _THIN_BORDER
