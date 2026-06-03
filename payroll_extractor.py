"""
PACE Payroll Validator — Payroll Extractor (Optimized)
=======================================================
Reads payroll Excel files using openpyxl read_only mode for speed.
Identifies pay-period columns with case/space-insensitive matching.
"""

from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
import pandas as pd

from logger_setup import setup_logger

logger = setup_logger()

# Canonical (normalized) column names we search for
_TARGET_START = "payperiodstart"
_TARGET_END = "payperiodend"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def extract_payroll_periods(
    file_paths: List[Path], woid: str
) -> Dict:
    """
    Extract unique payroll periods from all files and sheets.

    Returns
    -------
    Dict with keys:
        - ``periods``: List[Tuple[date, date]]
        - ``sheets_processed``: int
        - ``files_count``: int
    """
    all_periods: List[Tuple[date, date]] = []
    sheets_processed = 0

    for fpath in file_paths:
        try:
            wb = openpyxl.load_workbook(str(fpath), read_only=True, data_only=True)
        except Exception as exc:
            logger.error("WOID %s - Cannot open file %s: %s", woid, fpath.name, exc)
            continue

        try:
            for sheet_name in wb.sheetnames:
                result = _process_sheet_fast(wb, sheet_name, fpath.name, woid)
                if result is not None:
                    all_periods.extend(result)
                    sheets_processed += 1
        finally:
            wb.close()

    # Deduplicate
    unique_periods = list(set(all_periods))
    unique_periods.sort(key=lambda t: t[0])

    logger.info(
        "WOID %s - Extracted %d unique payroll period(s) from %d file(s), "
        "%d sheet(s) processed.",
        woid, len(unique_periods), len(file_paths), sheets_processed,
    )

    return {
        "periods": unique_periods,
        "sheets_processed": sheets_processed,
        "files_count": len(file_paths),
    }


# ---------------------------------------------------------------------------
# Fast sheet processing using openpyxl read_only
# ---------------------------------------------------------------------------

def _normalize(col_name: str) -> str:
    """Lowercase and strip all whitespace from a column name."""
    return "".join(col_name.lower().split())


def _process_sheet_fast(
    wb,
    sheet_name: str,
    file_name: str,
    woid: str,
) -> Optional[List[Tuple[date, date]]]:
    """
    Process a single sheet using openpyxl read_only mode (much faster
    than pandas for simple column extraction).
    """
    try:
        ws = wb[sheet_name]
    except Exception as exc:
        logger.warning("WOID %s - Could not read sheet '%s' in %s: %s",
                        woid, sheet_name, file_name, exc)
        return None

    # Read header row
    header_row = None
    start_col_idx = None
    end_col_idx = None

    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        header_row = row
        break

    if header_row is None:
        return None

    # Find target columns
    for idx, cell_val in enumerate(header_row):
        if cell_val is None:
            continue
        normalized = _normalize(str(cell_val))
        if normalized == _TARGET_START:
            start_col_idx = idx
        elif normalized == _TARGET_END:
            end_col_idx = idx

    if start_col_idx is None or end_col_idx is None:
        logger.debug("WOID %s - Sheet '%s' in %s missing pay-period columns, skipping.",
                      woid, sheet_name, file_name)
        return None

    # Extract data rows — only read the two columns we need
    seen = set()
    periods: List[Tuple[date, date]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Safely get values (row may be shorter than header)
        start_val = row[start_col_idx] if start_col_idx < len(row) else None
        end_val = row[end_col_idx] if end_col_idx < len(row) else None

        if start_val is None or end_val is None:
            continue

        # Skip blanks
        s_str = str(start_val).strip()
        e_str = str(end_val).strip()
        if not s_str or not e_str:
            continue

        # Deduplicate in-place
        key = (s_str, e_str)
        if key in seen:
            continue
        seen.add(key)

        s = _parse_date(start_val)
        e = _parse_date(end_val)
        if s is not None and e is not None:
            periods.append((s, e))
        else:
            logger.warning(
                "WOID %s - Unparseable dates in sheet '%s' of %s: start=%r, end=%r",
                woid, sheet_name, file_name, start_val, end_val,
            )

    if periods:
        logger.debug("WOID %s - Sheet '%s' in %s yielded %d period(s).",
                      woid, sheet_name, file_name, len(periods))

    return periods


def _parse_date(value) -> Optional[date]:
    """Best-effort date parsing from various formats."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    raw = str(value).strip()
    if not raw:
        return None

    # Common formats — try most likely first
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%Y-%m-%d %H:%M:%S.%f",
        "%m-%d-%Y",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue

    # Pandas fallback
    try:
        return pd.to_datetime(raw).date()
    except Exception:
        return None
